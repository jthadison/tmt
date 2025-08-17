"""
Transaction Export Module
Story 8.8 - Task 3: Build export functionality
"""
import csv
import json
import logging
from io import StringIO, BytesIO
from typing import List, Dict, Any, Optional
from datetime import datetime
from decimal import Decimal
from pathlib import Path
import asyncio

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from .transaction_manager import TransactionRecord
except ImportError:
    from transaction_manager import TransactionRecord

logger = logging.getLogger(__name__)


class TransactionExporter:
    """Handles exporting transactions to various formats"""
    
    def __init__(self):
        self.export_templates = {
            'standard': self._standard_template,
            'tax': self._tax_template,
            'detailed': self._detailed_template,
            'summary': self._summary_template
        }
        
    async def export_to_csv(self, 
                           transactions: List[TransactionRecord],
                           output_path: Optional[str] = None,
                           template: str = 'standard') -> str:
        """
        Export transactions to CSV format
        
        Args:
            transactions: List of transactions to export
            output_path: Optional file path to save CSV
            template: Export template to use
            
        Returns:
            CSV string or path to saved file
        """
        csv_buffer = StringIO()
        
        # Get template function
        template_func = self.export_templates.get(template, self._standard_template)
        headers, rows = template_func(transactions)
        
        # Write CSV
        writer = csv.DictWriter(csv_buffer, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
        
        csv_content = csv_buffer.getvalue()
        
        # Save to file if path provided
        if output_path:
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, 'w', newline='') as f:
                f.write(csv_content)
            logger.info(f"CSV exported to {output_path}")
            return output_path
            
        return csv_content
        
    async def export_to_excel(self, 
                            transactions: List[TransactionRecord],
                            output_path: str,
                            template: str = 'standard') -> str:
        """
        Export transactions to Excel format with formatting
        
        Args:
            transactions: List of transactions to export
            output_path: File path to save Excel
            template: Export template to use
            
        Returns:
            Path to saved file
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
            
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # Get template data
        template_func = self.export_templates.get(template, self._standard_template)
        headers, rows = template_func(transactions)
        
        # Add headers with formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        # Add data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Format numbers
                if header in ['P&L', 'Commission', 'Financing', 'Balance']:
                    cell.number_format = '#,##0.00'
                elif header == 'Units':
                    cell.number_format = '#,##0'
                    
                # Color code P&L
                if header == 'P&L' and value:
                    try:
                        pl_value = float(value)
                        if pl_value > 0:
                            cell.font = Font(color="008000")  # Green for profit
                        elif pl_value < 0:
                            cell.font = Font(color="FF0000")  # Red for loss
                    except (ValueError, TypeError):
                        pass
                        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
        # Add summary sheet if detailed template
        if template == 'detailed':
            summary_ws = wb.create_sheet("Summary")
            self._add_summary_sheet(summary_ws, transactions)
            
        # Save workbook
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info(f"Excel exported to {output_path}")
        return output_path
        
    async def export_to_pdf(self, 
                          transactions: List[TransactionRecord],
                          output_path: str,
                          template: str = 'standard') -> str:
        """
        Export transactions to PDF format
        
        Args:
            transactions: List of transactions to export
            output_path: File path to save PDF
            template: Export template to use
            
        Returns:
            Path to saved file
        """
        if not PDF_AVAILABLE:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
            
        # Create PDF document
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        doc = SimpleDocTemplate(output_path, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Add title
        title = Paragraph("Transaction Report", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 0.5*inch))
        
        # Add metadata
        metadata = Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
        story.append(metadata)
        story.append(Spacer(1, 0.25*inch))
        
        # Get template data
        template_func = self.export_templates.get(template, self._standard_template)
        headers, rows = template_func(transactions)
        
        # Create table data
        table_data = [headers]
        for row in rows[:100]:  # Limit to 100 rows for PDF
            table_data.append([row.get(h, '') for h in headers])
            
        # Create table
        table = Table(table_data)
        
        # Apply table style
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        # Add summary if more than 100 transactions
        if len(rows) > 100:
            story.append(Spacer(1, 0.25*inch))
            note = Paragraph(f"Note: Showing first 100 of {len(rows)} transactions", styles['Normal'])
            story.append(note)
            
        # Build PDF
        doc.build(story)
        logger.info(f"PDF exported to {output_path}")
        return output_path
        
    async def export_to_json(self, 
                           transactions: List[TransactionRecord],
                           output_path: Optional[str] = None) -> str:
        """
        Export transactions to JSON format
        
        Args:
            transactions: List of transactions to export
            output_path: Optional file path to save JSON
            
        Returns:
            JSON string or path to saved file
        """
        # Convert transactions to dictionaries
        data = {
            'export_date': datetime.now().isoformat(),
            'transaction_count': len(transactions),
            'transactions': [t.to_dict() for t in transactions]
        }
        
        json_content = json.dumps(data, indent=2, default=str)
        
        # Save to file if path provided
        if output_path:
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, 'w') as f:
                f.write(json_content)
            logger.info(f"JSON exported to {output_path}")
            return output_path
            
        return json_content
        
    def _standard_template(self, transactions: List[TransactionRecord]) -> tuple:
        """Standard export template"""
        headers = ['Date', 'Type', 'Instrument', 'Units', 'Price', 'P&L', 'Balance']
        rows = []
        
        for t in transactions:
            rows.append({
                'Date': t.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'Type': t.transaction_type,
                'Instrument': t.instrument or '',
                'Units': float(t.units),
                'Price': float(t.price),
                'P&L': float(t.pl),
                'Balance': float(t.account_balance)
            })
            
        return headers, rows
        
    def _tax_template(self, transactions: List[TransactionRecord]) -> tuple:
        """Tax reporting template"""
        headers = ['Date', 'Type', 'Instrument', 'Proceeds', 'Cost Basis', 'Gain/Loss', 'Commission']
        rows = []
        
        for t in transactions:
            if t.transaction_type in ['ORDER_FILL', 'TRADE_CLOSE']:
                rows.append({
                    'Date': t.timestamp.strftime('%Y-%m-%d'),
                    'Type': t.transaction_type,
                    'Instrument': t.instrument or '',
                    'Proceeds': float(abs(t.units) * t.price),
                    'Cost Basis': float(abs(t.units) * t.price - t.pl),
                    'Gain/Loss': float(t.pl),
                    'Commission': float(t.commission)
                })
                
        return headers, rows
        
    def _detailed_template(self, transactions: List[TransactionRecord]) -> tuple:
        """Detailed export template with all fields"""
        headers = ['Date', 'Type', 'Instrument', 'Units', 'Price', 'P&L', 
                  'Commission', 'Financing', 'Balance', 'Transaction ID', 
                  'Trade ID', 'Order ID', 'Reason']
        rows = []
        
        for t in transactions:
            rows.append({
                'Date': t.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'Type': t.transaction_type,
                'Instrument': t.instrument or '',
                'Units': float(t.units),
                'Price': float(t.price),
                'P&L': float(t.pl),
                'Commission': float(t.commission),
                'Financing': float(t.financing),
                'Balance': float(t.account_balance),
                'Transaction ID': t.transaction_id,
                'Trade ID': t.trade_id or '',
                'Order ID': t.order_id or '',
                'Reason': t.reason
            })
            
        return headers, rows
        
    def _summary_template(self, transactions: List[TransactionRecord]) -> tuple:
        """Summary template with aggregated data"""
        headers = ['Period', 'Total Trades', 'Total P&L', 'Win Rate', 'Avg Win', 'Avg Loss']
        
        # Calculate summary statistics
        total_pl = sum(t.pl for t in transactions)
        trades = [t for t in transactions if t.transaction_type in ['ORDER_FILL', 'TRADE_CLOSE']]
        wins = [t for t in trades if t.pl > 0]
        losses = [t for t in trades if t.pl < 0]
        
        win_rate = len(wins) / len(trades) * 100 if trades else 0
        avg_win = sum(t.pl for t in wins) / len(wins) if wins else Decimal('0')
        avg_loss = sum(t.pl for t in losses) / len(losses) if losses else Decimal('0')
        
        if transactions:
            period = f"{min(t.timestamp for t in transactions).date()} to {max(t.timestamp for t in transactions).date()}"
        else:
            period = "No transactions"
            
        rows = [{
            'Period': period,
            'Total Trades': len(trades),
            'Total P&L': float(total_pl),
            'Win Rate': f"{win_rate:.2f}%",
            'Avg Win': float(avg_win),
            'Avg Loss': float(avg_loss)
        }]
        
        return headers, rows
        
    def _add_summary_sheet(self, ws, transactions: List[TransactionRecord]):
        """Add summary statistics to Excel worksheet"""
        # Calculate statistics
        total_pl = sum(t.pl for t in transactions)
        total_commission = sum(t.commission for t in transactions)
        total_financing = sum(t.financing for t in transactions)
        trades = [t for t in transactions if t.transaction_type in ['ORDER_FILL', 'TRADE_CLOSE']]
        
        # Add summary data
        ws['A1'] = 'Summary Statistics'
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = 'Total Transactions:'
        ws['B3'] = len(transactions)
        
        ws['A4'] = 'Total P&L:'
        ws['B4'] = float(total_pl)
        ws['B4'].number_format = '#,##0.00'
        
        ws['A5'] = 'Total Commission:'
        ws['B5'] = float(total_commission)
        ws['B5'].number_format = '#,##0.00'
        
        ws['A6'] = 'Total Financing:'
        ws['B6'] = float(total_financing)
        ws['B6'].number_format = '#,##0.00'
        
        ws['A7'] = 'Total Trades:'
        ws['B7'] = len(trades)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15